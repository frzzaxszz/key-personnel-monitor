# -*- coding: utf-8 -*-
"""Excel 导入 / 导出"""
import io
import json

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from database import get_db
from models import Person, ImportLog

router = APIRouter(prefix="/api/import", tags=["import"])

HEADERS = ["姓名", "身份证号", "性别", "年龄", "手机号", "人员类别", "风险等级", "管控状态", "所属区域", "街道", "详细地址", "责任民警", "备注"]
REQUIRED_HEADERS = ["姓名", "身份证号"]
FIELD_DEFAULTS = {
    "性别": "男", "年龄": 0, "手机号": "", "人员类别": "涉稳人员",
    "风险等级": "低", "管控状态": "在控", "所属区域": "", "街道": "",
    "详细地址": "", "责任民警": "", "备注": "",
}


@router.get("/template")
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "重点人员导入"
    ws.append(HEADERS)
    ws.append(["张三", "110101199001011234", "男", 35, "13800138000", "涉稳人员", "中", "在控", "朝阳区", "和平街道", "阳光小区1栋101室", "张警官", "示例数据"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=persons_template.xlsx"})


@router.post("/excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not (file.filename or "").endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 文件")
    try:
        wb = load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception:
        raise HTTPException(400, "文件解析失败，请使用模板格式")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "文件为空")
    header = [str(c).strip() if c else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header) if h}
    missing_required = [h for h in REQUIRED_HEADERS if h not in idx]
    if missing_required:
        raise HTTPException(400, f"缺少必填列：{'、'.join(missing_required)}（请使用模板或补齐该列）")
    missing_optional = [h for h in FIELD_DEFAULTS if h not in idx]
    warnings = []
    if missing_optional:
        warnings.append(f"未识别的可选字段：{'、'.join(missing_optional)}，已使用系统默认值")

    def val(row, h):
        i = idx.get(h)
        if i is None or i >= len(row):
            return None
        v = row[i]
        return v if v is not None and str(v).strip() != "" else None

    success = 0
    failed = 0
    errors = []
    for rno, row in enumerate(rows[1:], start=2):
        if not row or not any(str(c).strip() if c else "" for c in row):
            continue
        name = val(row, "姓名")
        id_card = val(row, "身份证号")
        if not name or not id_card:
            failed += 1
            errors.append(f"第{rno}行：姓名/身份证号缺失")
            continue
        name = str(name).strip()
        id_card = str(id_card).strip()
        if db.query(Person).filter(Person.id_card == id_card).first():
            failed += 1
            errors.append(f"第{rno}行：身份证号 {id_card} 已存在")
            continue
        db.add(Person(
            name=name, id_card=id_card,
            gender=str(val(row, "性别") or FIELD_DEFAULTS["性别"]).strip() or "男",
            age=int(val(row, "年龄")) if isinstance(val(row, "年龄"), (int, float)) else 0,
            phone=str(val(row, "手机号") or "").strip(),
            category=str(val(row, "人员类别") or FIELD_DEFAULTS["人员类别"]).strip() or "涉稳人员",
            risk_level=str(val(row, "风险等级") or FIELD_DEFAULTS["风险等级"]).strip() or "低",
            control_status=str(val(row, "管控状态") or FIELD_DEFAULTS["管控状态"]).strip() or "在控",
            district=str(val(row, "所属区域") or "").strip(),
            street=str(val(row, "街道") or "").strip(),
            address=str(val(row, "详细地址") or "").strip(),
            manager=str(val(row, "责任民警") or "").strip(),
            notes=str(val(row, "备注") or "").strip(),
        ))
        success += 1
    db.add(ImportLog(filename=file.filename, total=success + failed, success=success, failed=failed, errors=json.dumps(errors[:100], ensure_ascii=False)))
    db.commit()
    return {"total": success + failed, "success": success, "failed": failed, "errors": errors[:100], "warnings": warnings}


@router.get("/export")
def export_excel(db: Session = Depends(get_db)):
    wb = Workbook()
    ws = wb.active
    ws.title = "重点人员"
    ws.append(HEADERS)
    for p in db.query(Person).all():
        ws.append([p.name, p.id_card, p.gender, p.age, p.phone, p.category, p.risk_level, p.control_status, p.district, p.street, p.address, p.manager, p.notes])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=persons_export.xlsx"})


@router.get("/logs")
def import_logs(db: Session = Depends(get_db)):
    logs = db.query(ImportLog).order_by(ImportLog.created_at.desc()).limit(20).all()
    return [{"id": l.id, "filename": l.filename, "total": l.total, "success": l.success, "failed": l.failed, "created_at": l.created_at} for l in logs]
